from tkinter import *
from tkinter import ttk
import datetime
import openpyxl
from PIL import Image
from openpyxl import Workbook

root = Tk()
#ikkunan mittaus
root.geometry('950x552')
#motto laitetaan missa kuva on.
root.iconbitmap()
#nimin kauppa
root.title('ML')

now = datetime.datetime.now()
date = now.strftime('%Y .%m .%d')

#===========date excel======
wb =Workbook()
ws =wb.active

ws.title ='customer'
ws['A1']= 'Full name'
ws['B1']= 'Number phone'
ws['C1'] ='Adress'
ws['D1'] = 'Total'
ws['E1'] ='Date buy'

wb.save('LM.xlsx')


def save():
	name = En_name.get()
	phone = En_phone.get()
	address = En_address.get()
	total = En_total.get()
	datebuy = En_date.get()

	excel = openpyxl.load_workbook('LM.xlsx')
	file = excel.active
	file.cell(column=1,row=file.max_row+1,value=name)
	file.cell(column=1,row=file.max_row,value=phone)
	file.cell(column=1,row=file.max_row,value=address)
	file.cell(column=1,row=file.max_row,value=total)
	file.cell(column=1,row=file.max_row,value=datebuy)

	excel.save('LM.xlsx')
#========= price ===========

menu ={
	0:['hiiri', 50],
	1:['hiirimato',20],
	2:['tietokoneHB',650],
	3:['kuulokkeet',45]
}

def bill():
	#global = voi käytetään tieto ilman def
	global En_name
	global En_address
	global En_phone
	global En_total
	global En_date
	



	Ib_image.place(x=950 ,y=438 )
	root.geometry('1250x552')
	F4 = Frame(root , bg ='#5f7161', width=250 , height = 434 , bd =2 , relief = GROOVE)
	F4.place(x=950, y=1) 
	L_name = Label (F4 , text= 'Ostajan nimi', bg = '#5f7161', fg = 'white')
	L_name.place(x=10,y =10)
	En_name = Entry(F4,width=23 ,font=('Tajawal',12 ),justify = CENTER)
	En_name.place(x=15 ,y=40)

	L_phone = Label (F4 , text= 'Puhelinnumero', bg = '#5f7161', fg = 'white')
	L_phone.place(x=10,y =70)
	En_phone = Entry(F4,width=23 ,font=('Tajawal',12 ),justify = CENTER)
	En_phone.place(x=15 ,y=100)


	L_address = Label (F4 , text= 'Osoite', bg = '#5f7161', fg = 'white')
	L_address.place(x=10,y =130)
	En_address = Entry(F4,width=23 ,font=('Tajawal',12 ),justify = CENTER)
	En_address.place(x=15 ,y=160)


	L_total= Label (F4 , text= 'Tili yhteensä', bg = '#5f7161', fg = 'white')
	L_total.place(x=10,y =190)
	En_total = Entry(F4,width=23 ,font=('Tajawal',12 ),justify = CENTER)
	En_total.place(x=15 ,y=210)


	L_date= Label (F4 , text = 'Ostopäivämäärä', bg = '#5f7161', fg = 'white')
	L_date.place(x=10,y =240)
	En_date = Entry(F4,width = 23 ,font = ('Tajawal',12 ),justify = CENTER)
	En_date.place(x=15 ,y=270)

	add_button = Button(F4,text ='Tallenna lasku',width = 31 ,cursor ='hand2',bg='#EDDBC0',command = save)
	add_button.place(x=10 ,y=310)

	add_button = Button(F4,text ='Tyhjennä kentät',width = 31 ,cursor ='hand2',bg='#EDDBC0', command = clear1)
	add_button.place(x=10 ,y=340)

	add_button = Button(F4,text ='',width = 31 ,cursor ='hand2',bg='#EDDBC0')
	add_button.place(x=10 ,y=370)

	add_button = Button(F4,text ='Poista lasku',width = 31 ,cursor ='hand2',bg ='#EDDBC0')
	add_button.place(x=10 ,y=400)


	total = 0
	for item in trv.get_children():
		trv.delete(item)

	for i in range(len(sb)):	
		if(int(sb[i].get())>0):
			price = int(sb[i].get())*menu[i][1]
			total = total + price
			myst = (str(menu[i][1]),str (sb[i].get()),str(price))
			trv.insert("",'end', iid = i , text =menu[i][0], values = myst)
	finall =total
	En_total.insert('1',str(finall) + '€')
	En_date.insert('1',str(date))


def clear():
	"""UUsi lasku"""
	for item in trv.get_children():
		trv.delete(item)
	En_name.delete('0',END)	
	En_address.delete('0',END)
	En_phone.delete('0',END)
	En_date.delete('0',END)

def clear1():
	"""Tyhjät kentät"""
	
	En_name.delete('0',END)	
	En_address.delete('0',END)
	En_phone.delete('0',END)
	En_total.delete('0',END)
	En_date.delete('0',END)



#======= [Frame[1]] =======

F1 = Frame(root ,bg ='silver',width=600,height=550)
F1.place(x = 1,y = 1)

#====== Image ========
img_menul = PhotoImage(file='img/hiiri.png')
img_menu2 = PhotoImage(file='img/hiirimato.png')
img_menu3 = PhotoImage(file='img/tietokoneHB.png')
img_menu4 = PhotoImage(file='img/kuulokkeet.png')

title = Label(F1, text='Tervetuloa ML-kauppaan',font=('Tajawal 13'),fg='white' ,bg ='#5F7161', width=70)

title.place(x = 0,y = 0)

menu1 = Button(F1 ,width = 88 , bg ='#EFEAD8' , bd = 1 ,relief = SOLID , cursor = 'hand2' ,height = 85 ,image = img_menul ,text='Hiiri' ,compound= TOP,)
menu1.place(x=30,y=45)

menu2 = Button(F1 ,width = 88 , bg ='#EFEAD8' , bd = 1 ,relief = SOLID , cursor = 'hand2' ,height = 85 ,image = img_menu2 ,text='Hiirimatto' ,compound= TOP,)
menu2.place(x=170,y=45)

menu3 = Button(F1 ,width = 88 , bg ='#EFEAD8' , bd = 1 ,relief = SOLID , cursor = 'hand2' ,height = 85 ,image = img_menu3 ,text='TietokoneHB' ,compound= TOP,)
menu3.place(x=310,y=45)

menu4 = Button(F1 ,width = 88 , bg ='#EFEAD8' , bd = 1 ,relief = SOLID , cursor = 'hand2' ,height = 85 ,image = img_menu4 ,text='kuulokkee' ,compound= TOP,)
menu4.place(x=450,y=45)


#Lisää kuvia
#x=(30 , 170 , 310 ,450),y=180
#x=(),y=320

#============ variable ==========
sb = []
font1 =( 'Times',12 ,'normal')

#tuotteiden määrä
sv1 = IntVar()
sv2 = IntVar()
sv3 = IntVar()
sv4 = IntVar()

sb1 = Spinbox(F1 , from_=0,to_=5,font = font1,width=10 ,textvariable =sv1)
sb1.place(x = 30 ,y = 140)
sb.append(sb1)


sb2 = Spinbox(F1 , from_=0,to_=5,font = font1,width=10 ,textvariable =sv2)
sb2.place(x = 170 ,y = 140)
sb.append(sb2)

sb3 = Spinbox(F1 , from_=0,to_=5,font = font1,width=10 ,textvariable =sv3)
sb3.place(x = 310 ,y = 140)
sb.append(sb3)

sb4 = Spinbox(F1 , from_=0,to_=5,font = font1,width=10 ,textvariable =sv4)
sb4.place(x = 450 ,y = 140)
sb.append(sb4)

#listaan lisa x =(30 ,170 ,310 , 450 ), y=275
#x=(sama) y=415

#=============Buttons=============
#40 min

b1 = Button(F1,text= 'Osta',fg ='white',font=('Tajawal 12'), width= 15 , bg ='#6D8B74',bd=1 ,relief = SOLID ,cursor ='hand2',height=1,command = bill)
b1.place(x=30,y=500)

b2 = Button(F1,text= 'Uusi lasku',fg ='white',font=('Tajawal 12'), width= 15 , bg ='#6D8B74',bd=1 ,relief = SOLID ,cursor ='hand2',height=1,command = clear)
b2.place(x=160,y=500)

b3 = Button(F1,text= 'Vuokra',fg ='white',font=('Tajawal 12'), width= 15 , bg ='#6D8B74',bd=1 ,relief = SOLID ,cursor ='hand2',height=1)
b3.place(x=290,y=500)

b4 = Button(F1,text= 'Poistu',fg ='white',font=('Tajawal 12'), width= 15 , bg ='#6D8B74',bd=1 ,relief = SOLID ,cursor ='hand2',height=1)
b4.place(x=420,y=500)

#=======Frame[2]=========
F2 = Frame(root, bg = 'Grey' ,width = 343 ,  height = 550)
F2.place(x = 604, y=1)

trv = ttk.Treeview(F2 ,selectmode ='browse')
trv.place(x = 1 ,y = 1 ,width =340,height = 550)

# Varaa paikkoja
trv['columns']=('1','2','3',)
trv.column('#0',width = 80 ,anchor = 'c')
trv.column('1',width = 50 ,anchor = 'c')
trv.column('2',width = 50 ,anchor = 'c')
trv.column('3',width = 60 ,anchor = 'c')

#lisää tautukoiden nimet
trv.heading('#0',text='Tuotteet',anchor ='c')
trv.heading('1',text='Hinta',anchor ='c')
trv.heading('2',text='Määrä',anchor ='c')
trv.heading('3',text='Yhteensä',anchor ='c')


# Lisää logon kuva

im_logo = PhotoImage(file='img/logo1.png')
#newsize = (300, 150)
#im_logo = im_lo.resize(newsize)
Ib_image = Label(root,image = im_logo)






root.mainloop()